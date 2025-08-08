import datetime as dt
import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side
from openpyxl.worksheet.worksheet import Worksheet

from .constants import TableSettings
from .exceptions import InvalidFilterColumnName, InvalidFilterValue


class ExcelTableFilter:
    def __init__(
            self,
            save_file_name: str = None,
            save_file_path: str = None,
            upload_file_path: str = None,
            needed_headers: list[str] = None,
            filter_column: str = None,
            filter_value: str | int | dt.date = None,
            table_header_indicator: str = TableSettings.TABLE_HEADER_INDICATOR,
    ):
        self.upload_file_path = upload_file_path
        self.save_file_name = save_file_name
        self.save_file_path = save_file_path

        self.needed_headers = needed_headers
        self.filter_column = filter_column
        self.filter_value = filter_value
        self.table_header_indicator = table_header_indicator

        self.work_sheet = None
        self.raw_data = []
        self.cleaned_data = []
        self.cleaned_headers = []
        self.filtered_data = []
        self.new_table = []

    def init_headers(self):
        if self.needed_headers is None:
            self.needed_headers = TableSettings.DEFAULT_RESULT_HEADERS

    def run_filter(self):
        self.init_headers()
        self.open_file()
        self.read_file()
        self.clean_data()
        self.filter_data()
        self.create_new_table()
        self.create_new_file()
        print('✅ Файл успешно обработан и сохранён.')

    def open_file(self) -> Worksheet:
        """Открывает Excel-файл по заданному пути и
        возвращает активный лист (Worksheet).
        """
        if not self.upload_file_path:
            raise ValueError('Путь к файлу не задан (upload_file_path)')
        self.work_sheet = load_workbook(self.upload_file_path).active
        return self.work_sheet

    def read_file(self) -> list[tuple]:
        """Считывает все строки из листа Excel,
        убирает полностью пустые строки.
        """
        for row in self.work_sheet.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                self.raw_data.append(row)
        return self.raw_data

    def clean_data(self) -> [tuple[list[str], list[list[str]]]]:
        """Находит строку-заголовок (по индикатору) и возвращает её отдельно,
        а также все строки после неё. Удаляет None из строк.
        """
        _data = [
            [item for item in row if item is not None]
            for row in self.raw_data
        ]
        _headers = []
        header_found = False
        for row in _data:
            if header_found:
                self.cleaned_data.append(row)
            if self.table_header_indicator in row:
                self.cleaned_headers = row
                header_found = True
        return self.cleaned_headers, self.cleaned_data

    def filter_data(self) -> list[tuple]:
        """Фильтрует строки по значению в выбранном столбце."""
        if self.filter_column is None:
            return []

        filter_col = self.filter_column.strip().lower()
        _headers = [h.strip().lower() for h in self.cleaned_headers]

        if filter_col not in _headers:
            raise InvalidFilterColumnName

        column_index = _headers.index(self.filter_column.lower())

        for row in self.cleaned_data:
            if len(row) <= column_index:
                continue
            cell = row[column_index]
            if cell is None:
                continue

            if isinstance(self.filter_value, str):
                if isinstance(cell, str):
                    cell_str = cell
                else:
                    cell_str = str(cell)
                if cell_str.strip().lower() == self.filter_value.strip().lower():
                    self.filtered_data.append(row)
            elif isinstance(self.filter_value, int):
                if cell == self.filter_value:
                    self.filtered_data.append(row)
            elif isinstance(self.filter_value, dt.date):
                try:
                    cell_date = dt.datetime.strptime(
                        cell,
                        TableSettings.CELL_DATE_FORMAT,
                    )
                    if cell_date.date() == self.filter_value.date():  # noqa
                        self.filtered_data.append(row)
                except (ValueError, Exception):
                    continue
        if not self.filtered_data:
            raise InvalidFilterValue(
                f'Значение \'{self.filter_value}\' '
                f'не найдено в колонке \'{self.filter_column}\''
            )
        return self.filtered_data

    def create_new_table(
            self,
    ) -> list[str]:
        """Оставляет только указанные столбцы (filter_headers)."""
        headers_indexes = [
            self.cleaned_headers.index(value)
            for value in self.needed_headers
        ]
        self.new_table = [
            [row[i] for i in headers_indexes]
            for row in self.filtered_data
        ]
        return self.new_table

    @staticmethod
    def add_borders(ws: Worksheet) -> Worksheet:
        """Добавляет тонкие чёрные границы."""
        border_side = Side(border_style='thin', color='000000')
        border = Border(
            left=border_side,
            right=border_side,
            top=border_side,
            bottom=border_side,
        )
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
        return ws

    def create_new_file(self) -> None:
        """Создаёт новый Excel-файл с заданными заголовками и данными.
        Применяет стилизацию границ к таблице.
        Сохраняет файл по указанному пути (или в текущей папке).
        """
        wb_new = Workbook()
        ws_new = wb_new.active
        ws_new.append(self.needed_headers)
        for row in self.new_table:
            ws_new.append(row)
        self.add_borders(ws_new)

        if self.save_file_path and self.save_file_name:
            full_path = os.path.join(self.save_file_path, self.save_file_name)
        elif self.save_file_name:
            full_path = os.path.abspath(self.save_file_name)
        else:
            raise ValueError(
                'Ошибка формирования пути сохранения файла: '
                'save_file_name не задан.'
            )
        wb_new.save(full_path)


excel_service = ExcelTableFilter()
